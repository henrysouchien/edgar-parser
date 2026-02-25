import re
import os
import json
import time
import logging
import urllib.error
import urllib.request
import requests
from datetime import timedelta, date, datetime, UTC
from collections import defaultdict
from .config import (
    HEADERS,
    ANTHROPIC_MODEL_8K,
    ENABLE_8K_LLM_FALLBACK,
    MAX_8K_HTML_BYTES,
    MAX_8K_HTML_BYTES_OPENAI,
    OPENAI_API_KEY,
    OPENAI_MODEL_8K,
)
from .http_client import get as http_get
from .utils import lookup_cik_from_ticker, parse_date

logger = logging.getLogger(__name__)

# === LLM API telemetry ===
# Pricing per million tokens
_CLAUDE_INPUT_COST_PER_M = 3.00
_CLAUDE_OUTPUT_COST_PER_M = 15.00
_OPENAI_INPUT_COST_PER_M = 2.50
_OPENAI_OUTPUT_COST_PER_M = 10.00


def log_llm_api(
    ticker,
    year,
    quarter,
    model,
    input_tokens,
    output_tokens,
    duration_sec,
    status,
    error_msg=None,
    provider="anthropic",
):
    """Log LLM API call metrics to usage_logs/claude_api_log.jsonl."""
    if provider == "openai":
        input_rate = _OPENAI_INPUT_COST_PER_M
        output_rate = _OPENAI_OUTPUT_COST_PER_M
    else:
        provider = "anthropic"
        input_rate = _CLAUDE_INPUT_COST_PER_M
        output_rate = _CLAUDE_OUTPUT_COST_PER_M
    cost_usd = (input_tokens / 1_000_000) * input_rate + (output_tokens / 1_000_000) * output_rate
    record = {
        "timestamp": datetime.now(UTC).isoformat(),
        "ticker": ticker,
        "year": year,
        "quarter": quarter,
        "model": model,
        "input_tokens": input_tokens,
        "output_tokens": output_tokens,
        "duration_sec": round(duration_sec, 2),
        "cost_usd": round(cost_usd, 4),
        "status": status,
        "provider": provider,
    }
    if error_msg:
        record["error"] = error_msg

    os.makedirs("usage_logs", exist_ok=True)
    with open("usage_logs/claude_api_log.jsonl", "a") as f:
        f.write(json.dumps(record) + "\n")


def log_claude_api(
    ticker, year, quarter, model, input_tokens, output_tokens, duration_sec, status, error_msg=None
):
    """Backward-compatible wrapper for Anthropic log records."""
    return log_llm_api(
        ticker,
        year,
        quarter,
        model,
        input_tokens,
        output_tokens,
        duration_sec,
        status,
        error_msg,
        provider="anthropic",
    )


# TODO: n_limit=8 means only the 8 most recent Item 2.02 8-Ks are fetched.
# Requests for older quarters may fail if the correct 8-K falls outside this
# window. Consider increasing or making caller-configurable if needed.
def fetch_recent_8k_accessions(cik, headers, n_limit=8):
    url = f"https://data.sec.gov/submissions/CIK{cik.zfill(10)}.json"
    r = http_get(url, headers=headers)
    r.raise_for_status()
    filings = r.json()["filings"]["recent"]

    results = []
    fallback_701 = []
    for i, form in enumerate(filings["form"]):
        if form != "8-K":
            continue
        # SEC submissions JSON has an "items" field for 8-Ks -- comma-separated item codes
        # Item 2.02 = "Results of Operations" (standard earnings release)
        # Item 7.01 = "Regulation FD Disclosure" (some companies file earnings here instead)
        items_str = filings.get("items", [""] * len(filings["form"]))[i] or ""
        if "2.02" not in items_str and "7.01" not in items_str:
            continue
        entry = {
            "accession": filings["accessionNumber"][i],
            "report_date": filings["reportDate"][i],
            "filing_date": filings["filingDate"][i],
            "items": items_str,
        }
        if "2.02" in items_str:
            results.append(entry)
            if len(results) >= n_limit:
                break
        elif "7.01" in items_str:
            fallback_701.append(entry)
    if results:
        return results
    # Fallback: use 7.01 if no 2.02 filings were found
    return fallback_701[:n_limit]


def fetch_8k_exhibit(cik, accession, headers):
    acc_nodash = accession.replace("-", "")
    index_url = f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{acc_nodash}/index.json"
    try:
        r = http_get(index_url, headers=headers, rate_limited=True)
        r.raise_for_status()
    except requests.RequestException:
        return None, None
    directory = r.json().get("directory", {}).get("item", [])

    exhibit_name = None

    # Priority 1: type field -- prefer EX-99.1 over EX-99.2, HTML only (skip PDFs)
    ex99_by_type = []
    for item in directory:
        item_type = item.get("type", "").upper()
        name_lower = item.get("name", "").lower()
        if item_type.startswith("EX-99") and name_lower.endswith((".htm", ".html")):
            ex99_by_type.append((item_type, item["name"]))
    # Sort numerically: extract the suffix after "EX-99." so EX-99.2 < EX-99.10
    def _ex99_sort_key(pair):
        m = re.search(r"EX-99\.?(\d+)", pair[0], re.IGNORECASE)
        return int(m.group(1)) if m else 999
    ex99_by_type.sort(key=_ex99_sort_key)
    if ex99_by_type:
        exhibit_name = ex99_by_type[0][1]

    # Priority 2: filename pattern (HTML only)
    if not exhibit_name:
        for item in directory:
            name_lower = item.get("name", "").lower()
            if ("ex99" in name_lower or "ex-99" in name_lower) and name_lower.endswith((".htm", ".html")):
                exhibit_name = item["name"]
                break

    # Priority 3: largest .htm that isn't the 8-K cover or index
    # Only exclude specific known non-exhibit files -- don't exclude by "8k" substring
    # because some real exhibits may contain "8k" in their filename
    if not exhibit_name:
        skip_names = set()
        for item in directory:
            name_lower = item.get("name", "").lower()
            item_type = item.get("type", "").upper()
            # Skip the primary 8-K document and index files
            if item_type in ("8-K", "8-K/A"):
                skip_names.add(item["name"])
            if name_lower.endswith(("-index.htm", "-index.html", "index.htm", "index.html")):
                skip_names.add(item["name"])
        htm_files = [
            (item["name"], int(item.get("size", 0) or 0))
            for item in directory
            if item["name"].lower().endswith((".htm", ".html"))
            and item["name"] not in skip_names
        ]
        if not htm_files:
            return None, None
        exhibit_name = max(htm_files, key=lambda x: x[1])[0]

    exhibit_url = f"https://www.sec.gov/Archives/edgar/data/{int(cik)}/{acc_nodash}/{exhibit_name}"
    try:
        r = http_get(exhibit_url, headers=headers, rate_limited=True)
        r.raise_for_status()
    except requests.RequestException:
        return None, None
    return r.text, exhibit_url


def find_8k_for_period(cik, headers, year, quarter, metadata_only=False):
    from .tools import (
        enrich_10k_accessions_with_fiscal_year,
        fetch_recent_10q_10k_accessions,
        label_10q_accessions,
    )

    # 1. Get Item 2.02 8-Ks
    eight_k_list = fetch_recent_8k_accessions(cik, headers)
    if not eight_k_list:
        return None, None, None, None

    # 2. Compute expected period-end using existing fiscal calendar logic
    #    Reuse the same approach as the 10-Q/10-K pipeline:
    #    - fetch_recent_10q_10k_accessions() gives us 10-Q and 10-K lists
    #    - label_10q_accessions() assigns quarters to 10-Qs using FY-end from 10-Ks
    #    - For the requested year/quarter, look up the corresponding period-end date
    accessions_10q, accessions_10k = fetch_recent_10q_10k_accessions(cik, headers)
    try:
        accessions_10q = label_10q_accessions(accessions_10q, accessions_10k)
    except ValueError:
        # No 10-Ks with valid dates (e.g., recent IPO) — can't determine fiscal calendar
        accessions_10q = []
    accessions_10k = enrich_10k_accessions_with_fiscal_year(accessions_10k)

    # Look up expected period-end from the fiscal calendar
    # Primary: exact match from labeled 10-Qs or 10-Ks
    # Fallback: derive expected period-end using the same fiscal-year-end
    # alignment logic the main pipeline uses (via labeled 10-Qs)
    expected_period_end = None

    if quarter == 4:
        # Q4 period-end = fiscal year-end (from 10-K reportDate)
        for k in accessions_10k:
            if k.get("year") == year:
                expected_period_end = parse_date(k.get("report_date"))
                break
        if not expected_period_end:
            # No 10-K for requested year -- use closest 10-K's FY-end, shift to requested year
            # (FY-end month/day is consistent year-to-year for the vast majority of companies)
            for k in accessions_10k:
                if k.get("fiscal_year_end"):
                    expected_period_end = k["fiscal_year_end"].replace(year=year)
                    break
    else:
        # Q1-Q3: find the exact labeled 10-Q for the requested period
        target_label = f"{quarter}Q{str(year)[-2:]}"
        for q in accessions_10q:
            if q.get("label") == target_label:
                expected_period_end = parse_date(q.get("report_date"))
                break

        # Compute FY-end info upfront (needed by both ref-10-Q shift and midpoint fallback)
        fy_end_dates = [k["fiscal_year_end"] for k in accessions_10k if k.get("fiscal_year_end")]
        fy_end_consistent = len(set((d.month, d.day) for d in fy_end_dates)) <= 1 if fy_end_dates else True
        fye_target = None
        for k in accessions_10k:
            if k.get("year") == year and k.get("fiscal_year_end"):
                fye_target = k["fiscal_year_end"]
                break
        if not fye_target and fy_end_dates:
            fye_target = fy_end_dates[0].replace(year=year)

        if not expected_period_end:
            # No 10-Q for this quarter yet -- use pipeline-style fiscal-year-end alignment:
            # 1) Identify the target fiscal year-end (from 10-Ks) for the requested year.
            # 2) Find a prior 10-Q for the same quarter with a matching fiscal-year-end month/day.
            # 3) Compute the day offset between that 10-Q's report_date and its FY-end,
            #    then apply the same offset to the target FY-end.
            if fye_target and fy_end_consistent:
                target_quarter = f"Q{quarter}"
                # Find the MOST RECENT reference 10-Q with the same quarter and FY-end month/day
                ref_candidates = []
                for q in accessions_10q:
                    if q.get("quarter") != target_quarter or not q.get("report_date"):
                        continue
                    q_fye = q.get("fiscal_year_end")
                    if q_fye and (q_fye.month, q_fye.day) == (fye_target.month, fye_target.day):
                        ref_candidates.append(q)
                ref_q = None
                if ref_candidates:
                    ref_q = max(
                        ref_candidates,
                        key=lambda q: parse_date(q.get("report_date")) or date.min,
                    )
                if ref_q:
                    ref_end = parse_date(ref_q.get("report_date"))
                    ref_fye = ref_q.get("fiscal_year_end")
                    if ref_end and ref_fye:
                        delta_days = (ref_fye - ref_end).days
                        expected_period_end = fye_target - timedelta(days=delta_days)

        # Last-resort fallback: use label_10q_accessions day-diff midpoints
        # (same ranges the pipeline uses to assign quarters)
        # Q1: 250-300 days before FY-end → midpoint 275
        # Q2: 160-200 → midpoint 180
        # Q3: 70-120 → midpoint 95
        # Only used when FY-end is consistent (or known for the exact year),
        # since the midpoints assume a stable fiscal calendar.
        if not expected_period_end and fye_target and fy_end_consistent:
            DAY_DIFF_MIDPOINTS = {1: 275, 2: 180, 3: 95}
            midpoint = DAY_DIFF_MIDPOINTS.get(quarter)
            if midpoint:
                expected_period_end = fye_target - timedelta(days=midpoint)

    if not expected_period_end:
        return None, None, None, None

    # 3. Find the EARLIEST Item 2.02/7.01 8-K filed AFTER the expected period-end
    #    (within a 150-day window) with lightweight period validation on the exhibit content.
    #    Sort ascending by filing_date so we pick the closest 8-K to the period-end,
    #    not the most recent one (the SEC list is newest-first).
    #    150 days accommodates late Q4 filers (10-K deadline is 60-90 days post FY-end,
    #    but the 8-K earnings release can occasionally trail that for smaller companies).
    MAX_8K_WINDOW_DAYS = 150
    candidates = []
    for entry in eight_k_list:
        filing_date = parse_date(entry.get("filing_date"))
        if filing_date and filing_date >= expected_period_end:
            gap = (filing_date - expected_period_end).days
            if gap <= MAX_8K_WINDOW_DAYS:
                candidates.append((filing_date, entry))
    candidates.sort(key=lambda x: x[0])  # earliest first

    for _, entry in candidates:
        if metadata_only:
            # Skip HTML download — return metadata only (used by get_filings)
            return entry, None, None, str(expected_period_end)
        html, exhibit_url = fetch_8k_exhibit(cik, entry["accession"], headers)
        if not html:
            continue
        # Lightweight validation: if we can parse a period-end date from the
        # exhibit, check it matches. If we can't parse one (regex doesn't cover
        # all phrasings), accept the 8-K anyway -- the filing_date window already
        # constrains the search to a reasonable range.
        exhibit_date = _extract_period_end_from_html(html, expected_period_end)
        if exhibit_date and abs((exhibit_date - expected_period_end).days) > 15:
            continue  # period mismatch -- skip, try next candidate
        period_end_str = str(exhibit_date or expected_period_end)
        return entry, html, exhibit_url, period_end_str

    return None, None, None, None


# Regex to extract period-end dates from exhibit HTML
# Handles: "Three months ended December 31, 2025", "Quarter ended Dec. 31, 2025",
#           "Fiscal quarter ended March 29, 2025", "13 weeks ended January 1, 2026",
#           "Three months ended December 31, 2025 (unaudited)", etc.
MONTH = (
    r"(?:January|February|March|April|May|June|July|August|September|October|November|December"
    r"|Jan\.?|Feb\.?|Mar\.?|Apr\.?|May\.?|Jun\.?|Jul\.?|Aug\.?|Sep\.?|Oct\.?|Nov\.?|Dec\.?)"
)
PERIOD_DATE_PATTERN = re.compile(
    r"(?:for\s+the\s+)?"
    r"(?:fiscal\s+)?"  # optional "fiscal" prefix
    r"(?:"
    r"(?:three|six|nine|twelve|thirteen|fourteen|52|53)\s+(?:months|weeks)\s+ended"
    r"|quarter\s+ended"
    r")\s+"
    r"(" + MONTH + r"\s+\d{1,2},?\s+\d{4})"
    r"(?:\s*\(unaudited\))?",  # optional trailing "(unaudited)"
    re.IGNORECASE,
)


def _extract_period_end_from_html(html, expected_period_end):
    """Extract a period-end date from the exhibit HTML. Returns closest match or None."""
    matches = PERIOD_DATE_PATTERN.findall(html)
    if not matches:
        return None
    parsed = []
    for m in matches:
        dt = parse_date(m)
        if dt:
            parsed.append(dt)
    if not parsed:
        return None
    # Prefer the date closest to the expected period-end
    return min(parsed, key=lambda d: abs((d - expected_period_end).days))


def _strip_html_attrs(html_text):
    """Strip all attributes from HTML tags to reduce token count.
    Keeps the tag structure (table, tr, td, th, span, etc.) but removes
    style, class, id, width, and other attributes that bloat SEC filings.

    NOTE: This also removes colspan/rowspan attributes, which can affect table
    alignment for filings that use merged cells. In practice, SEC earnings
    press releases rarely use complex cell spans, and Claude handles the
    simplified structure well. If extraction quality degrades for a specific
    company, consider preserving colspan/rowspan by switching to a more
    selective regex, e.g.:
        re.sub(r'(?!\\s(?:colspan|rowspan)=)\\s[a-z-]+=("[^"]*"|...)', '', tag)
    """
    return re.sub(r"<(\/?[a-zA-Z][a-zA-Z0-9]*)\b[^>]*\/?>", r"<\1>", html_text)


def _truncate_8k_html(html, max_bytes):
    if len(html.encode("utf-8")) <= max_bytes:
        return html

    tables = re.findall(
        r"(.{0,200}<table>.*?</table>.{0,200})",
        html,
        flags=re.DOTALL | re.IGNORECASE,
    )
    if tables:
        html = "\n\n".join(tables)

    html_bytes = html.encode("utf-8")
    if len(html_bytes) <= max_bytes:
        return html

    trimmed = html_bytes[:max_bytes].decode("utf-8", errors="ignore")
    return trimmed.rsplit(" ", 1)[0] if " " in trimmed else trimmed


def _preprocess_8k_html(html_content):
    html = re.sub(
        r"<(style|script|head)[^>]*>.*?</\1>",
        "",
        html_content,
        flags=re.DOTALL | re.IGNORECASE,
    )
    html = _strip_html_attrs(html)
    html = re.sub(r"<(td|th|span|div|p)>\s*</\1>", "", html, flags=re.IGNORECASE)
    html = re.sub(r"\s{2,}", " ", html)
    return _truncate_8k_html(html, MAX_8K_HTML_BYTES)


def _status_code_from_error(exc):
    status_code = getattr(exc, "status_code", None)
    if isinstance(status_code, int):
        return status_code
    response = getattr(exc, "response", None)
    response_status = getattr(response, "status_code", None)
    return response_status if isinstance(response_status, int) else None


def _is_anthropic_server_error(exc, anthropic_errors=None):
    anthropic_errors = anthropic_errors or {}
    overloaded = anthropic_errors.get("OverloadedError")
    internal = anthropic_errors.get("InternalServerError")
    api_status = anthropic_errors.get("APIStatusError")
    if isinstance(overloaded, type) and isinstance(exc, overloaded):
        return True
    if isinstance(internal, type) and isinstance(exc, internal):
        return True
    if isinstance(api_status, type) and isinstance(exc, api_status):
        status_code = _status_code_from_error(exc)
        return isinstance(status_code, int) and status_code >= 500
    if isinstance(exc, urllib.error.HTTPError):
        return exc.code >= 500
    name = exc.__class__.__name__
    if name in {"OverloadedError", "InternalServerError"}:
        return True
    if name == "APIStatusError":
        status_code = _status_code_from_error(exc)
        return isinstance(status_code, int) and status_code >= 500
    return False


def _is_retriable_error(exc, anthropic_errors=None):
    if isinstance(exc, (ValueError, TypeError, KeyError)):
        return False
    if _is_anthropic_server_error(exc, anthropic_errors):
        return True

    anthropic_errors = anthropic_errors or {}
    for name in ("RateLimitError", "APIConnectionError", "AuthenticationError"):
        error_type = anthropic_errors.get(name)
        if isinstance(error_type, type) and isinstance(exc, error_type):
            return True
    if isinstance(exc, urllib.error.HTTPError):
        return exc.code >= 500

    return exc.__class__.__name__ in {
        "RateLimitError",
        "APIConnectionError",
        "AuthenticationError",
    }


def _safe_usage_int(value):
    try:
        return max(int(value or 0), 0)
    except (TypeError, ValueError):
        return 0


def _call_openai(prompt, html, model):
    """Call OpenAI chat completions API. Returns (response_text, input_tokens, output_tokens)."""
    if not OPENAI_API_KEY:
        raise ValueError("OPENAI_API_KEY is not set")

    html_payload = _truncate_8k_html(html, MAX_8K_HTML_BYTES_OPENAI)
    payload = {
        "model": model,
        "temperature": 0,
        "max_completion_tokens": 8192,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": prompt},
            {"role": "user", "content": html_payload},
        ],
    }
    req = urllib.request.Request(
        "https://api.openai.com/v1/chat/completions",
        data=json.dumps(payload).encode("utf-8"),
        method="POST",
        headers={
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            body = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail_raw = exc.read().decode("utf-8", errors="ignore")
        detail = detail_raw
        if detail_raw:
            try:
                parsed_detail = json.loads(detail_raw)
                err = parsed_detail.get("error") if isinstance(parsed_detail, dict) else None
                if isinstance(err, dict):
                    detail = err.get("message") or detail_raw
            except json.JSONDecodeError:
                pass
        raise ValueError(f"OpenAI API HTTP {exc.code}: {detail or exc.reason}") from exc
    except urllib.error.URLError as exc:
        raise ValueError(f"OpenAI API connection error: {exc.reason}") from exc

    choices = body.get("choices") if isinstance(body, dict) else None
    if not isinstance(choices, list) or not choices:
        raise ValueError("OpenAI API returned no choices")

    message = choices[0].get("message") if isinstance(choices[0], dict) else None
    content = message.get("content") if isinstance(message, dict) else None
    if isinstance(content, str):
        raw_text = content.strip()
    elif isinstance(content, list):
        parts = []
        for item in content:
            if isinstance(item, dict):
                text_part = item.get("text")
                if isinstance(text_part, str):
                    parts.append(text_part)
            elif isinstance(item, str):
                parts.append(item)
        raw_text = "".join(parts).strip()
    else:
        raise ValueError("OpenAI API returned unexpected content format")
    if not raw_text:
        raise ValueError("OpenAI API returned empty response content")

    usage = body.get("usage") if isinstance(body, dict) else {}
    usage_map = usage if isinstance(usage, dict) else {}
    input_tokens = _safe_usage_int(usage_map.get("prompt_tokens"))
    output_tokens = _safe_usage_int(usage_map.get("completion_tokens"))
    return raw_text, input_tokens, output_tokens


def _parse_raw_facts(raw_text, ticker, year, quarter, provider_name):
    cleaned = raw_text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```\w*\n?", "", cleaned)
        cleaned = re.sub(r"\n?```$", "", cleaned)

    try:
        raw_facts = json.loads(cleaned)
    except json.JSONDecodeError as exc:
        bracket_match = re.search(r"\[.*\]", cleaned, re.DOTALL)
        if bracket_match:
            try:
                raw_facts = json.loads(bracket_match.group())
            except json.JSONDecodeError as bracket_exc:
                raise ValueError(
                    f"{provider_name} returned invalid JSON for {ticker} Q{quarter} {year}: {bracket_exc}. "
                    f"First 200 chars: {cleaned[:200]}"
                ) from bracket_exc
        else:
            raise ValueError(
                f"{provider_name} returned invalid JSON for {ticker} Q{quarter} {year}: {exc}. "
                f"First 200 chars: {cleaned[:200]}"
            ) from exc

    if not isinstance(raw_facts, list):
        raise ValueError(
            f"{provider_name} returned {type(raw_facts).__name__} instead of list for {ticker} Q{quarter} {year}. "
            f"First 200 chars: {cleaned[:200]}"
        )
    return raw_facts


def extract_facts_from_8k(html_content, ticker, year, quarter, full_year_mode):
    anthropic_available = False
    anthropic_error_types = {}
    Anthropic = None
    try:
        import anthropic as anthropic_module

        Anthropic = anthropic_module.Anthropic
        anthropic_available = True
        anthropic_error_types = {
            "OverloadedError": getattr(anthropic_module, "OverloadedError", None),
            "InternalServerError": getattr(anthropic_module, "InternalServerError", None),
            "RateLimitError": getattr(anthropic_module, "RateLimitError", None),
            "APIConnectionError": getattr(anthropic_module, "APIConnectionError", None),
            "AuthenticationError": getattr(anthropic_module, "AuthenticationError", None),
            "APIStatusError": getattr(anthropic_module, "APIStatusError", None),
        }
    except ImportError:
        anthropic_available = False

    html = _preprocess_8k_html(html_content)

    # 2. Build prompt
    period = f"FY {year}" if full_year_mode else f"Q{quarter} {year}"
    prompt = f"""This is {ticker}'s earnings press release for {period}.

Extract all financial line items from every table (quarterly, YTD,
full-year, and balance sheet) as year-over-year comparisons -- the
actual numbers, not the change -- with their labels.

Output a JSON array where each item has:
- "tag": a descriptive label for the line item. If the row label is generic
  (e.g., "Diluted", "Basic"), prepend the section or table header to
  disambiguate (e.g., "Earnings per share, diluted", "Shares outstanding, basic")
- "current": the most recent period's number for that table
- "prior": the prior-year comparison number for that same period
- "date_type": must be exactly one of these string values:
  - "Q" -- single quarter (e.g., "three months ended"). Balance sheet items ("as of" a date) ALWAYS use "Q" regardless of the date.
  - "YTD" -- year-to-date or cumulative (e.g., "nine months ended")
  - "FY" -- full fiscal year (e.g., "twelve months ended" or annual)
- "scale": integer power of 10 representing the unit scale of the values.
  Use 0 for actual values (e.g., per-share data, ratios, counts),
  3 for thousands, 6 for millions, 9 for billions.
  Look for cues like "in millions, except per share data" in table
  headers or filing text. Per-share items (EPS, dividends per share)
  are always scale 0.

Numbers in parentheses are negative. Strip commas but don't apply any
scale factor. Use null for missing values.

Output ONLY the JSON array, no other text."""

    def _parse_and_postprocess(
        provider_name, model, input_tokens, output_tokens, duration, raw_text, provider
    ):
        try:
            raw_facts = _parse_raw_facts(raw_text, ticker, year, quarter, provider_name)
        except ValueError as exc:
            status = "invalid_type" if "instead of list" in str(exc) else "invalid_json"
            log_llm_api(
                ticker,
                year,
                quarter,
                model,
                input_tokens,
                output_tokens,
                duration,
                status,
                str(exc),
                provider=provider,
            )
            raise
        log_llm_api(
            ticker,
            year,
            quarter,
            model,
            input_tokens,
            output_tokens,
            duration,
            "success",
            provider=provider,
        )
        return _postprocess_facts(raw_facts)

    if anthropic_available:
        client = Anthropic()  # uses ANTHROPIC_API_KEY env var
        anthropic_error = None

        for attempt in range(2):
            start_time = time.time()
            try:
                response = client.messages.create(
                    model=ANTHROPIC_MODEL_8K,
                    max_tokens=8192,
                    messages=[
                        {
                            "role": "user",
                            "content": [
                                {"type": "text", "text": prompt},
                                {"type": "text", "text": html},
                            ],
                        }
                    ],
                )
                duration = time.time() - start_time
                input_tokens = getattr(response.usage, "input_tokens", 0)
                output_tokens = getattr(response.usage, "output_tokens", 0)

                if not response.content:
                    err = ValueError(
                        f"Anthropic API returned empty response for {ticker} Q{quarter} {year}"
                    )
                    log_claude_api(
                        ticker,
                        year,
                        quarter,
                        ANTHROPIC_MODEL_8K,
                        input_tokens,
                        output_tokens,
                        duration,
                        "empty_response",
                        str(err),
                    )
                    raise err

                content_block = response.content[0] if response.content else None
                raw_text = getattr(content_block, "text", "") if content_block else ""
                if not isinstance(raw_text, str) or not raw_text.strip():
                    err = ValueError(
                        f"Anthropic API returned empty response for {ticker} Q{quarter} {year}"
                    )
                    log_claude_api(
                        ticker,
                        year,
                        quarter,
                        ANTHROPIC_MODEL_8K,
                        input_tokens,
                        output_tokens,
                        duration,
                        "empty_response",
                        str(err),
                    )
                    raise err

                return _parse_and_postprocess(
                    "Claude",
                    ANTHROPIC_MODEL_8K,
                    input_tokens,
                    output_tokens,
                    duration,
                    raw_text,
                    provider="anthropic",
                )
            except ValueError as exc:
                anthropic_error = exc
                break
            except Exception as exc:
                anthropic_error = exc
                duration = time.time() - start_time
                log_claude_api(
                    ticker,
                    year,
                    quarter,
                    ANTHROPIC_MODEL_8K,
                    0,
                    0,
                    duration,
                    "error",
                    str(exc),
                )
                if attempt == 0 and _is_anthropic_server_error(exc, anthropic_error_types):
                    time.sleep(2)
                    continue
                break

        if anthropic_error and not _is_retriable_error(anthropic_error, anthropic_error_types):
            raise anthropic_error
    else:
        anthropic_error = ImportError("anthropic package is unavailable")

    fallback_enabled = ENABLE_8K_LLM_FALLBACK
    has_openai_key = bool(OPENAI_API_KEY)
    if not fallback_enabled or not has_openai_key:
        if anthropic_available and anthropic_error is not None:
            if not fallback_enabled:
                logger.warning(
                    "8-K LLM fallback disabled; re-raising Anthropic error for %s Q%s %s",
                    ticker,
                    quarter,
                    year,
                )
            elif not has_openai_key:
                logger.warning(
                    "8-K LLM fallback requested but OPENAI_API_KEY is missing for %s Q%s %s",
                    ticker,
                    quarter,
                    year,
                )
            raise ValueError(
                f"Anthropic API error for {ticker} Q{quarter} {year}: "
                f"{type(anthropic_error).__name__}: {anthropic_error}"
            ) from anthropic_error
        raise ValueError("Neither anthropic nor OpenAI fallback is available for 8-K extraction")

    start_time = time.time()
    try:
        raw_text, input_tokens, output_tokens = _call_openai(prompt, html, OPENAI_MODEL_8K)
    except ValueError as openai_error:
        duration = time.time() - start_time
        log_llm_api(
            ticker,
            year,
            quarter,
            OPENAI_MODEL_8K,
            0,
            0,
            duration,
            "error",
            str(openai_error),
            provider="openai",
        )
        if anthropic_available and anthropic_error is not None:
            raise ValueError(
                f"Anthropic and OpenAI fallback both failed for {ticker} Q{quarter} {year}: "
                f"anthropic={type(anthropic_error).__name__}: {anthropic_error}; "
                f"openai={openai_error}"
            ) from openai_error
        raise ValueError(
            f"OpenAI fallback failed for {ticker} Q{quarter} {year}: {openai_error}"
        ) from openai_error

    duration = time.time() - start_time
    try:
        return _parse_and_postprocess(
            "OpenAI",
            OPENAI_MODEL_8K,
            input_tokens,
            output_tokens,
            duration,
            raw_text,
            provider="openai",
        )
    except ValueError as openai_error:
        if anthropic_available and anthropic_error is not None:
            raise ValueError(
                f"Anthropic and OpenAI fallback both failed for {ticker} Q{quarter} {year}: "
                f"anthropic={type(anthropic_error).__name__}: {anthropic_error}; "
                f"openai={openai_error}"
            ) from openai_error
        raise ValueError(
            f"OpenAI fallback failed for {ticker} Q{quarter} {year}: {openai_error}"
        ) from openai_error


def _coerce_numeric(val):
    """Ensure value is numeric. Handle strings with commas/parens from Claude output.

    NOTE: Values like "15%", "N/M", "—" will return None. The prompt instructs Claude
    to output raw numbers without % signs, so this should be rare. If margin metrics
    with % become common, consider stripping % and dividing by 100.
    """
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        s = val.strip().replace(",", "")
        # Handle common placeholder dashes
        if s in ("—", "-", "–", "N/A", "N/M", ""):
            return None
        if s.startswith("(") and s.endswith(")"):
            s = "-" + s[1:-1]
        try:
            return int(s) if "." not in s else float(s)
        except ValueError:
            return None
    return None


VALID_DATE_TYPES = {"Q", "YTD", "FY"}
SCALE_LABELS = {-2: "hundredths", 0: "units", 3: "thousands", 6: "millions", 9: "billions"}


def _postprocess_facts(raw_facts):
    """Map Claude's {tag, current, prior, date_type} to full EdgarFact schema."""
    facts = []
    for item in raw_facts:
        current = _coerce_numeric(item.get("current"))
        prior = _coerce_numeric(item.get("prior"))
        raw_dt = item.get("date_type")
        date_type = raw_dt.upper().strip() if isinstance(raw_dt, str) else None
        date_type = date_type if date_type in VALID_DATE_TYPES else None
        raw_scale = item.get("scale")
        try:
            scale_int = int(float(raw_scale)) if raw_scale is not None else None
        except (ValueError, TypeError):
            scale_int = None
        scale_int = scale_int if scale_int in (-2, 0, 3, 6, 9) else None
        scale = SCALE_LABELS.get(scale_int) if scale_int is not None else None
        facts.append(
            {
                "tag": item.get("tag"),
                "date_type": date_type,
                "presentation_role": None,
                "current_period_value": current,
                "prior_period_value": prior,
                "visual_current_value": current,
                "visual_prior_value": prior,
                "end_current": None,
                "end_prior": None,
                "axis_segment": None,
                "axis_geo": None,
                "collision_flag": 0,  # computed below
                "scale": scale,
            }
        )

    # Collision detection: same prior → multiple distinct currents
    # Skip prior values that are 0 or null (too common, would spike false collisions)
    prior_to_currents = defaultdict(set)
    for f in facts:
        p = f["prior_period_value"]
        c = f["current_period_value"]
        if p is not None and p != 0 and c is not None:
            prior_to_currents[p].add(c)

    colliding_priors = {p for p, currents in prior_to_currents.items() if len(currents) > 1}
    for f in facts:
        if f["prior_period_value"] in colliding_priors:
            f["collision_flag"] = 1

    return facts


def get_financials_from_8k(ticker, year, quarter, full_year_mode=False, use_cache=True):
    # === 1. Check cache ===
    cache_dir = "exports"
    cache_filename = (
        f"{ticker}_FY{str(year)[-2:]}_8k_financials.json"
        if full_year_mode
        else f"{ticker}_{quarter}Q{str(year)[-2:]}_8k_financials.json"
    )
    cache_path = os.path.join(cache_dir, cache_filename)

    # NOTE: app.py also checks this cache at the endpoint level for /api/financials.
    # This is intentionally redundant so CLI calls (which bypass Flask) also get caching.
    if use_cache and os.path.exists(cache_path):
        with open(cache_path, "r") as f:
            cached = json.load(f)
        facts = cached.get("facts", [])
        if facts and "scale" in facts[0]:
            return cached

    # === 2. CIK lookup ===
    cik = lookup_cik_from_ticker(ticker)
    if not cik:
        return {"status": "error", "message": f"Could not find CIK for {ticker}"}

    # === 3. Find 8-K for period ===
    entry, html, exhibit_url, expected_period_end = find_8k_for_period(cik, HEADERS, year, quarter)
    if not entry:
        return {"status": "error", "message": f"No Item 2.02 8-K found for {ticker} Q{quarter} {year}"}

    # Detect value scale from HTML (informational only)
    scale_match = re.search(r"in\s+(thousands|millions|billions)", html, re.IGNORECASE)
    value_scale = scale_match.group(1).lower() if scale_match else "unknown"

    # === 4. Extract facts via Claude ===
    try:
        facts = extract_facts_from_8k(html, ticker, year, quarter, full_year_mode)
    except ValueError as e:
        return {"status": "error", "message": str(e)}

    result = {
        "status": "success",
        "metadata": {
            "ticker": ticker,
            "year": year,
            "quarter": quarter,
            "full_year_mode": full_year_mode,
            "total_facts": len(facts),
            "source": {
                "filing_type": "8-K",
                "period_end": expected_period_end or entry.get("filing_date"),
                "url": exhibit_url,
                "value_scale": value_scale,
                "cik": cik,
                "accession": entry["accession"],
            },
        },
        "facts": facts,
    }

    # === 5. Save to cache ===
    os.makedirs(cache_dir, exist_ok=True)
    with open(cache_path, "w") as f:
        json.dump(result, f)

    return result


def write_8k_facts_to_excel(facts, ticker, year, quarter, full_year_mode, excel_file):
    """Write 8-K extracted facts to Raw_data sheet, matching pipeline Excel format."""
    import openpyxl

    wb = openpyxl.load_workbook(excel_file, keep_vba=True)
    sheet = wb["Raw_data"]

    # Clear existing data
    for row in sheet["A2:F5000"]:
        for cell in row:
            cell.value = None

    # Write header for scale column
    sheet["F1"] = "Scale"

    # Write facts
    for i, fact in enumerate(facts):
        row_num = i + 2
        sheet.cell(row=row_num, column=1, value=fact.get("tag"))
        sheet.cell(
            row=row_num,
            column=2,
            value=fact.get("visual_current_value", fact.get("current_period_value")),
        )
        sheet.cell(
            row=row_num,
            column=3,
            value=fact.get("visual_prior_value", fact.get("prior_period_value")),
        )
        sheet.cell(row=row_num, column=4, value=fact.get("presentation_role", ""))
        sheet.cell(row=row_num, column=5, value=fact.get("collision_flag", 0))
        sheet.cell(row=row_num, column=6, value=fact.get("scale"))

    # Write metadata
    sheet["G1"] = "Ticker"
    sheet["H1"] = ticker
    sheet["G2"] = "Year"
    sheet["H2"] = year
    sheet["G3"] = "Quarter"
    sheet["H3"] = quarter
    sheet["G4"] = "Full Year Mode"
    sheet["H4"] = str(full_year_mode)

    wb.save(excel_file)
